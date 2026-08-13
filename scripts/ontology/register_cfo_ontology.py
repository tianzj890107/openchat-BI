# -*- coding: utf-8 -*-
"""
把 CFO「订-发-收-回」故事线的 8 张 T_CFO_* 表完整登记进
ChatBI业务元数据_硕磐财务管理.xlsx 的全部 9 个本体页签:

  术语 / 业务规则 / 活动 / 业务对象 / 逻辑实体 / 业务属性 / 实体关系 / 指标 / 指标维度矩阵

续号(基于现状):术语 T000050+ / 规则 R035+ / 活动 A012 / 业务对象 BO0050 /
逻辑实体 LE00053+ / 业务属性 AT00359+ / 实体关系 ER105+ / 指标 M049+。
脚本可重复执行:执行前先按编码前缀清除上一次写入的 CFO 行,保证幂等。
"""

import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from bi_agent.paths import SPREADSHEETS_DIR

XLSX = PROJECT_ROOT / SPREADSHEETS_DIR / "ChatBI业务元数据_硕磐财务管理.xlsx"
CUR, PRE = "2026-05", "2026-04"

# ---- 8 张表的列规格: (中文名, 物理列, 数据类型, 定义, 是否主键) ----------
TABLES = [
    ("LE00053", "T_CFO_StoryIndex", "CFO故事线索引", "cfo_story_index",
     "CFO「订-发-收-回」四条故事线的索引,标注每条故事线对应的取数表、口径单位与本月/上月期间。", [
        ("故事线编号", "FStoryNo", "INT", "故事线序号(1订货/2增收/3现金/4盈利)。", "Yes"),
        ("分析方向", "FDirection", "NVARCHAR(20)", "故事线方向:订货/增收/现金/盈利。", "No"),
        ("核心问题", "FQuestion", "NVARCHAR(500)", "该故事线 CFO 关心的核心问题。", "No"),
        ("对应表", "FTable", "NVARCHAR(200)", "回答该故事线的主物理表(可多张)。", "No"),
        ("金额单位", "FUnit", "NVARCHAR(20)", "金额口径单位,统一为万元。", "No"),
        ("本月期间", "FCurMonth", "NVARCHAR(20)", "本月期间(YYYY-MM)。", "No"),
        ("上月期间", "FPreMonth", "NVARCHAR(20)", "上月对照期间(YYYY-MM)。", "No"),
     ]),
    ("LE00054", "T_CFO_OrderQuality", "订单质量分析", "cfo_order_quality",
     "故事线1:按月×维度(合计/区域/渠道)度量订单增长质量,含环比、低毛利占比、高折扣占比、可转收入比例。", [
        ("期间", "FMonth", "NVARCHAR(20)", "统计月份(YYYY-MM)。", "Yes"),
        ("维度类型", "FDimType", "NVARCHAR(20)", "维度类型:合计/区域/渠道。", "Yes"),
        ("维度值", "FDimValue", "NVARCHAR(50)", "维度取值(如华东、大客户KA)。", "Yes"),
        ("订单额", "FOrderAmtWan", "DECIMAL(18,2)", "本月订单额(万元)。", "No"),
        ("上月订单额", "FPreOrderAmtWan", "DECIMAL(18,2)", "上月订单额(万元)。", "No"),
        ("环比增长率", "FMoMRatePct", "DECIMAL(9,2)", "订单额环比增长率(%)。", "No"),
        ("低毛利订单占比", "FLowMarginRatioPct", "DECIMAL(9,2)", "低毛利订单金额占比(%)。", "No"),
        ("高折扣订单占比", "FHighDiscRatioPct", "DECIMAL(9,2)", "高折扣(折扣≥15%)订单占比(%)。", "No"),
        ("平均毛利率", "FAvgGrossMarginPct", "DECIMAL(9,2)", "订单平均毛利率(%)。", "No"),
        ("平均折扣率", "FAvgDiscountPct", "DECIMAL(9,2)", "订单平均折扣率(%)。", "No"),
        ("可转收入比例", "FConvertibleRatioPct", "DECIMAL(9,2)", "预计可在本月/下月转为收入的订单比例(%)。", "No"),
        ("备注", "FNote", "NVARCHAR(500)", "质量风险说明。", "No"),
     ]),
    ("LE00055", "T_CFO_HighRiskOrder", "高风险订单清单", "cfo_high_risk_order",
     "故事线1后续动作:需销售/供应链/财务联合跟进的高风险订单明细,分高金额未发货/低毛利高折扣/交期临近库存不足三类。", [
        ("订单号", "FOrderNo", "NVARCHAR(40)", "销售订单号。", "Yes"),
        ("客户", "FCustomer", "NVARCHAR(80)", "下单客户名称。", "No"),
        ("区域", "FRegion", "NVARCHAR(40)", "所属销售区域。", "No"),
        ("产品线", "FProductLine", "NVARCHAR(40)", "订单主产品线。", "No"),
        ("订单额", "FOrderAmtWan", "DECIMAL(18,2)", "订单金额(万元)。", "No"),
        ("毛利率", "FGrossMarginPct", "DECIMAL(9,2)", "订单毛利率(%)。", "No"),
        ("折扣率", "FDiscountPct", "DECIMAL(9,2)", "订单折扣率(%)。", "No"),
        ("订单状态", "FStatus", "NVARCHAR(30)", "未发货/已发货未签收/已签收未开票。", "No"),
        ("风险类型", "FRiskType", "NVARCHAR(40)", "高金额未发货/低毛利高折扣/交期临近库存不足。", "No"),
        ("库存缺口数量", "FStockGapQty", "DECIMAL(18,2)", "交期类风险的关键物料库存缺口数量。", "No"),
        ("交期", "FDueDate", "NVARCHAR(20)", "客户要求交期。", "No"),
        ("跟进部门", "FOwnerDept", "NVARCHAR(60)", "联合跟进部门(销售/供应链/财务)。", "No"),
        ("备注", "FNote", "NVARCHAR(500)", "风险说明。", "No"),
     ]),
    ("LE00056", "T_CFO_OrderToRevenue", "订单转收入卡点", "cfo_order_to_revenue",
     "故事线2:按环节(未发货/已发货未签收/已签收未开票/已开票确认收入)拆解订单转收入的转化卡点与收入确认影响。", [
        ("期间", "FMonth", "NVARCHAR(20)", "统计月份(YYYY-MM)。", "Yes"),
        ("环节", "FStage", "NVARCHAR(30)", "转化环节名称。", "Yes"),
        ("环节序号", "FStageOrder", "INT", "环节顺序(1→4)。", "No"),
        ("订单额", "FOrderAmtWan", "DECIMAL(18,2)", "该环节滞留订单额(万元)。", "No"),
        ("订单数", "FOrderCount", "INT", "该环节订单笔数。", "No"),
        ("占比", "FRatioPct", "DECIMAL(9,2)", "占本月新增订单比例(%)。", "No"),
        ("主要产品线", "FMainProductLine", "NVARCHAR(40)", "该环节集中的主要产品线。", "No"),
        ("主要区域", "FMainRegion", "NVARCHAR(40)", "该环节集中的主要区域。", "No"),
        ("根因", "FRootCause", "NVARCHAR(200)", "该环节滞留的根本原因。", "No"),
        ("收入确认影响", "FRevenueImpactWan", "DECIMAL(18,2)", "该环节滞留导致的本月收入确认影响(万元)。", "No"),
     ]),
    ("LE00057", "T_CFO_RevenueToCash", "收入转现金", "cfo_revenue_to_cash",
     "故事线3:月度收入与回款对比,经营现金转化率、应收净增加、逾期应收与现金安全评估。", [
        ("期间", "FMonth", "NVARCHAR(20)", "统计月份(YYYY-MM)。", "Yes"),
        ("收入", "FRevenueWan", "DECIMAL(18,2)", "本月收入(万元)。", "No"),
        ("上月收入", "FPreRevenueWan", "DECIMAL(18,2)", "上月收入(万元)。", "No"),
        ("收入环比", "FRevenueMoMPct", "DECIMAL(9,2)", "收入环比增长率(%)。", "No"),
        ("回款", "FCollectionWan", "DECIMAL(18,2)", "本月回款(万元)。", "No"),
        ("上月回款", "FPreCollectionWan", "DECIMAL(18,2)", "上月回款(万元)。", "No"),
        ("回款环比", "FCollectionMoMPct", "DECIMAL(9,2)", "回款环比增长率(%)。", "No"),
        ("现金转化率", "FCashConvRatePct", "DECIMAL(9,2)", "经营现金转化率=回款/收入(%)。", "No"),
        ("上月现金转化率", "FPreCashConvRatePct", "DECIMAL(9,2)", "上月经营现金转化率(%)。", "No"),
        ("应收余额", "FAROpenWan", "DECIMAL(18,2)", "期末应收账款余额(万元)。", "No"),
        ("应收净增加", "FARDeltaWan", "DECIMAL(18,2)", "本月应收净增加(万元)≈收入-回款。", "No"),
        ("逾期应收", "FOverdueAmtWan", "DECIMAL(18,2)", "逾期应收金额(万元)。", "No"),
        ("现金安全评估", "FCashSafety", "NVARCHAR(30)", "现金流安全评估:安全/短期可控/承压。", "No"),
        ("备注", "FNote", "NVARCHAR(500)", "口径与风险说明。", "No"),
     ]),
    ("LE00058", "T_CFO_KeyCustomerAR", "重点客户回款跟踪", "cfo_key_customer_ar",
     "故事线3:重点客户应收余额、账期变化、逾期与对应收净增加的贡献,支撑高风险应收督办。", [
        ("客户", "FCustomer", "NVARCHAR(80)", "客户名称。", "Yes"),
        ("客户分层", "FCustomerTier", "NVARCHAR(20)", "大客户KA/重点/一般。", "No"),
        ("区域", "FRegion", "NVARCHAR(40)", "客户所属区域。", "No"),
        ("应收余额", "FAROpenWan", "DECIMAL(18,2)", "客户应收余额(万元)。", "No"),
        ("应收增加", "FARDeltaWan", "DECIMAL(18,2)", "本月该客户应收增加(万元)。", "No"),
        ("当前账期", "FAvgTermDays", "INT", "当前平均账期(天)。", "No"),
        ("上月账期", "FPreTermDays", "INT", "上月平均账期(天)。", "No"),
        ("逾期金额", "FOverdueAmtWan", "DECIMAL(18,2)", "逾期金额(万元)。", "No"),
        ("最大逾期天数", "FMaxOverdueDays", "INT", "最大逾期天数。", "No"),
        ("应收贡献占比", "FContribRatioPct", "DECIMAL(9,2)", "占全公司应收净增加比例(%)。", "No"),
        ("督办动作", "FAction", "NVARCHAR(200)", "高风险应收督办动作。", "No"),
     ]),
    ("LE00059", "T_CFO_ProfitBridge", "利润桥分析", "cfo_profit_bridge",
     "故事线4:净利润环比的驱动分解(规模/价格端/成本端/费用端),定位增长未带来利润的主因。", [
        ("期间", "FMonth", "NVARCHAR(20)", "统计月份(YYYY-MM)。", "Yes"),
        ("收入", "FRevenueWan", "DECIMAL(18,2)", "本月收入(万元)。", "No"),
        ("上月收入", "FPreRevenueWan", "DECIMAL(18,2)", "上月收入(万元)。", "No"),
        ("毛利率", "FGrossMarginPct", "DECIMAL(9,2)", "本月毛利率(%)。", "No"),
        ("上月毛利率", "FPreGrossMarginPct", "DECIMAL(9,2)", "上月毛利率(%)。", "No"),
        ("费用", "FExpenseWan", "DECIMAL(18,2)", "本月费用(万元)。", "No"),
        ("上月费用", "FPreExpenseWan", "DECIMAL(18,2)", "上月费用(万元)。", "No"),
        ("净利润", "FNetProfitWan", "DECIMAL(18,2)", "本月净利润(万元)。", "No"),
        ("上月净利润", "FPreNetProfitWan", "DECIMAL(18,2)", "上月净利润(万元)。", "No"),
        ("净利环比", "FNetProfitMoMPct", "DECIMAL(9,2)", "净利润环比增长率(%)。", "No"),
        ("规模贡献", "FVolumeEffectWan", "DECIMAL(18,2)", "规模/销量对利润的贡献(+,万元)。", "No"),
        ("价格端影响", "FPriceEffectWan", "DECIMAL(18,2)", "价格端(折扣↑/低毛利客户↑)影响(-,万元)。", "No"),
        ("成本端影响", "FCostEffectWan", "DECIMAL(18,2)", "成本端(原料/制造↑)影响(-,万元)。", "No"),
        ("费用端影响", "FExpenseEffectWan", "DECIMAL(18,2)", "费用端(费用投入↑)影响(-,万元)。", "No"),
        ("主因", "FMainDriver", "NVARCHAR(30)", "利润下滑主因(价格端/成本端/费用端)。", "No"),
        ("备注", "FNote", "NVARCHAR(500)", "利润桥勾稽说明。", "No"),
     ]),
    ("LE00060", "T_CFO_LowMarginList", "低毛利高折扣清单", "cfo_low_margin_list",
     "故事线4:按产品线/客户/区域列出最拉低利润的低毛利、高折扣业务,支撑价格复核与利润改善督办。", [
        ("维度类型", "FDimType", "NVARCHAR(20)", "维度类型:产品线/客户/区域。", "Yes"),
        ("维度值", "FDimValue", "NVARCHAR(80)", "维度取值。", "Yes"),
        ("收入", "FRevenueWan", "DECIMAL(18,2)", "该维度收入(万元)。", "No"),
        ("毛利率", "FGrossMarginPct", "DECIMAL(9,2)", "毛利率(%)。", "No"),
        ("平均折扣率", "FAvgDiscountPct", "DECIMAL(9,2)", "平均折扣率(%)。", "No"),
        ("利润拖累", "FProfitDragWan", "DECIMAL(18,2)", "对利润的负向拖累(万元)。", "No"),
        ("改善动作", "FAction", "NVARCHAR(200)", "价格复核/利润改善督办动作。", "No"),
     ]),
]

BO_CODE = "BO0050"
BO_NAME = "CFO经营分析"
BO_TABLES = "; ".join(t[1] for t in TABLES)

TERMS = [  # 术语编码 名称 别名 英文 定义 分类 解释部门
    ("T000050", "高质量增长", "有效增长、健康增长", "High-quality Growth",
     "订单/收入增长同时维持或改善毛利、折扣与回款结构;若增长伴随低毛利与高折扣占比上升,则为带质量风险的增长。", "经营分析", "财务管理部"),
    ("T000051", "订单质量", "订单结构", "Order Quality",
     "从毛利率、折扣率、可转收入比例等角度衡量订单增长的有效性。", "经营分析", "财务管理部"),
    ("T000052", "订单转收入转化率", "订单收入转化、订转收", "Order-to-Revenue Conversion",
     "新增订单中在本月/下月完成发货、签收、开票并确认为收入的比例;卡点分布在未发货、已发货未签收、已签收未开票。", "经营分析", "财务管理部"),
    ("T000053", "经营现金转化率", "现金转化率、收现率", "Cash Conversion Rate",
     "本月回款 / 本月收入,衡量收入真正转化为现金的程度;低于收入增速说明收入滞留在应收。", "经营分析", "财务管理部"),
    ("T000054", "应收账期", "账期、信用期", "Receivable Term Days",
     "客户从开票到回款的平均天数;账期拉长与逾期增加会加大资金周转压力。", "经营分析", "财务管理部"),
    ("T000055", "利润桥", "利润桥分析、净利驱动分解", "Profit Bridge",
     "把净利润环比变化拆解为规模、价格端、成本端、费用端四类驱动,定位利润变动主因。", "经营分析", "财务管理部"),
    ("T000056", "价格端影响", "价格因素、折扣侵蚀", "Price Effect",
     "由折扣上升、低毛利客户占比提高导致的利润侵蚀,是利润桥中价格相关的负向驱动。", "经营分析", "财务管理部"),
    ("T000057", "低毛利订单", "低毛利业务", "Low-margin Order",
     "毛利率显著低于公司均值的订单,占比上升会拉低整体利润。", "经营分析", "财务管理部"),
    ("T000058", "高折扣订单", "高折扣业务", "High-discount Order",
     "折扣率达到或超过阈值(≥15%)的订单,折扣过高侵蚀毛利。", "经营分析", "财务管理部"),
    ("T000059", "高风险订单", "风险订单清单", "High-risk Order",
     "需联合跟进的订单,含高金额未发货、低毛利高折扣、交期临近库存不足三类。", "经营分析", "财务管理部"),
    ("T000060", "重点客户回款", "大客户回款、应收督办", "Key Customer Collection",
     "对应收净增加贡献最大的重点客户的回款跟踪与高风险应收督办。", "经营分析", "财务管理部"),
]

ACTIVITIES = [  # 活动编码 名称 描述 上游 角色 操作对象编号 操作对象名称
    ("A012", "CFO经营故事线分析",
     "围绕订-发-收-回与盈利四个方向,对订单质量、订单转收入卡点、收入转现金、利润桥进行分析并触发联合跟进与督办。",
     "A010", "CFO/财务/销售/供应链", BO_CODE, BO_NAME),
]

RULES = [  # 规则编号 规则 所属活动 活动名称 规则分类 解释部门
    ("R035", "订单环比增长率=(本月订单额-上月订单额)/上月订单额;增长同时低毛利与高折扣占比上升判定为带质量风险的增长", "A012", "CFO经营故事线分析", "分析规则", "财务管理部"),
    ("R036", "高风险订单按三类识别:高金额未发货、低毛利高折扣(折扣≥15%且毛利率偏低)、交期临近且库存缺口>0", "A012", "CFO经营故事线分析", "预警规则", "财务管理部"),
    ("R037", "订单转收入按环节拆解(未发货/已发货未签收/已签收未开票/已开票确认收入),各环节金额合计等于本月新增订单额", "A012", "CFO经营故事线分析", "稽核规则", "财务管理部"),
    ("R038", "经营现金转化率=回款/收入;收入增速持续高于回款增速且转化率下降时触发高风险应收督办", "A012", "CFO经营故事线分析", "预警规则", "财务管理部"),
    ("R039", "应收净增加≈本月收入-本月回款;重点客户按对应收净增加的贡献占比排序跟踪", "A012", "CFO经营故事线分析", "分析规则", "财务管理部"),
    ("R040", "利润桥勾稽:上月净利+规模贡献+价格端+成本端+费用端=本月净利;负向最大项即利润下滑主因", "A012", "CFO经营故事线分析", "稽核规则", "财务管理部"),
    ("R041", "低毛利/高折扣业务按产品线、客户、区域列示利润拖累,触发价格复核与利润改善督办", "A012", "CFO经营故事线分析", "分析规则", "财务管理部"),
]

RELATIONS = [  # 关系编号 源编号 源名 目标编号 目标名 类型 描述 外键
    ("ER105", "LE00053", "CFO故事线索引", "LE00054", "订单质量分析", "1:N", "故事线1索引指向订单质量分析表", "FTable"),
    ("ER106", "LE00053", "CFO故事线索引", "LE00056", "订单转收入卡点", "1:N", "故事线2索引指向订单转收入卡点表", "FTable"),
    ("ER107", "LE00053", "CFO故事线索引", "LE00057", "收入转现金", "1:N", "故事线3索引指向收入转现金表", "FTable"),
    ("ER108", "LE00053", "CFO故事线索引", "LE00059", "利润桥分析", "1:N", "故事线4索引指向利润桥分析表", "FTable"),
    ("ER109", "LE00054", "订单质量分析", "LE00055", "高风险订单清单", "1:N", "订单质量分析下钻到高风险订单清单(同区域/产品线)", "FRegion/FProductLine"),
    ("ER110", "LE00057", "收入转现金", "LE00058", "重点客户回款跟踪", "1:N", "收入转现金按重点客户下钻应收贡献", "FMonth"),
    ("ER111", "LE00059", "利润桥分析", "LE00060", "低毛利高折扣清单", "1:N", "利润桥价格端下钻到低毛利/高折扣清单", "FMonth"),
    ("ER112", "LE00055", "高风险订单清单", "LE00002", "产品线", "N:1", "高风险订单关联产品线", "FProductLine->FName"),
    ("ER113", "LE00058", "重点客户回款跟踪", "LE00009", "客户", "N:1", "重点客户回款关联客户主数据", "FCustomer->FName"),
]

# 指标: (编码,名称,别名,英文,定义,业务公式,口径,类型,技术公式,表,聚合列,聚合,连接,筛选,维度,版本,说明)
METRICS = [
    ("M049", "本月订单额", "订单额、新签订单额", "Monthly Order Amount",
     "本月订单总金额(万元)", "SUM(FOrderAmtWan)", "按月/区域/渠道", 2, None,
     "T_CFO_OrderQuality", "FOrderAmtWan", "SUM", None, "FDimType='合计'", None, "v1", "CFO故事线1"),
    ("M050", "订单环比增长率", "订单MoM", "Order MoM Growth",
     "订单额环比增长率(%)", "(本月订单额-上月订单额)/上月订单额", "按月", 1, None,
     "T_CFO_OrderQuality", "FMoMRatePct", "AVG", None, "FDimType='合计'", None, "v1", "CFO故事线1"),
    ("M051", "低毛利订单占比", "低毛利占比", "Low-margin Order Ratio",
     "低毛利订单金额占比(%)", "AVG(FLowMarginRatioPct)", "按月/区域/渠道", 2, None,
     "T_CFO_OrderQuality", "FLowMarginRatioPct", "AVG", None, "FDimType='合计'", None, "v1", "CFO故事线1"),
    ("M052", "高折扣订单占比", "高折扣占比", "High-discount Order Ratio",
     "高折扣订单金额占比(%)", "AVG(FHighDiscRatioPct)", "按月/区域/渠道", 2, None,
     "T_CFO_OrderQuality", "FHighDiscRatioPct", "AVG", None, "FDimType='合计'", None, "v1", "CFO故事线1"),
    ("M053", "订单可转收入比例", "可转收入率", "Order Convertible Ratio",
     "预计可转为收入的订单比例(%)", "AVG(FConvertibleRatioPct)", "按月", 2, None,
     "T_CFO_OrderQuality", "FConvertibleRatioPct", "AVG", None, "FDimType='合计'", None, "v1", "CFO故事线1"),
    ("M054", "订单转收入滞留金额", "卡点金额、滞留订单额", "Stuck Order Amount",
     "未确认收入环节滞留的订单额(万元)", "SUM(FOrderAmtWan) WHERE 环节<>已开票", "按月/环节", 2, None,
     "T_CFO_OrderToRevenue", "FOrderAmtWan", "SUM", None, "FStage<>'已开票确认收入'", None, "v1", "CFO故事线2"),
    ("M055", "经营现金转化率", "现金转化率、收现率", "Cash Conversion Rate",
     "回款/收入(%)", "FCollectionWan/FRevenueWan", "按月", 1, None,
     "T_CFO_RevenueToCash", "FCashConvRatePct", "AVG", None, None, None, "v1", "CFO故事线3"),
    ("M056", "应收净增加", "应收增加额", "AR Net Increase",
     "本月应收净增加(万元)", "SUM(FARDeltaWan)", "按月", 2, None,
     "T_CFO_RevenueToCash", "FARDeltaWan", "SUM", None, None, None, "v1", "CFO故事线3"),
    ("M057", "逾期应收金额", "逾期金额", "Overdue AR Amount",
     "逾期应收金额(万元)", "SUM(FOverdueAmtWan)", "按月/客户", 2, None,
     "T_CFO_RevenueToCash", "FOverdueAmtWan", "SUM", None, None, None, "v1", "CFO故事线3"),
    ("M058", "重点客户应收贡献占比", "客户应收贡献", "Key Customer AR Contribution",
     "重点客户对应收净增加的贡献占比(%)", "SUM(FContribRatioPct)", "按客户", 2, None,
     "T_CFO_KeyCustomerAR", "FContribRatioPct", "SUM", None, "FCustomerTier='大客户KA'", None, "v1", "CFO故事线3"),
    ("M059", "净利润环比增长率", "净利MoM", "Net Profit MoM Growth",
     "净利润环比增长率(%)", "(本月净利-上月净利)/上月净利", "按月", 1, None,
     "T_CFO_ProfitBridge", "FNetProfitMoMPct", "AVG", None, None, None, "v1", "CFO故事线4"),
    ("M060", "价格端利润影响", "价格端影响、折扣侵蚀额", "Price Effect on Profit",
     "价格端对净利润的影响(万元,负向)", "SUM(FPriceEffectWan)", "按月", 2, None,
     "T_CFO_ProfitBridge", "FPriceEffectWan", "SUM", None, None, None, "v1", "CFO故事线4"),
    ("M061", "低毛利业务利润拖累", "利润拖累", "Low-margin Profit Drag",
     "低毛利/高折扣业务对利润的拖累(万元)", "SUM(FProfitDragWan)", "按产品线/客户/区域", 2, None,
     "T_CFO_LowMarginList", "FProfitDragWan", "SUM", None, None, None, "v1", "CFO故事线4"),
]

# 指标维度矩阵 dims 顺序(与表头一致, 16 个):
# 时间(期间) 事业部(组织) 产品线 物料 客户 供应商 会计科目 管报分类 库龄桶 预警等级
# 状态类型 预算版本 滚动轮次 内外部标识 领用人 领用部门
def dimrow(code, name, time=0, org=0, pl=0, mat=0, cust=0, sup=0, acc=0,
           mgmt=0, age=0, warn=0, status=0, bud=0, roll=0, inout=0, pp=0, pd=0):
    y = lambda x: "Y" if x else None
    return (code, name, y(time), y(org), y(pl), y(mat), y(cust), y(sup),
            y(acc), y(mgmt), y(age), y(warn), y(status), y(bud), y(roll),
            y(inout), y(pp), y(pd))

DIMROWS = [
    dimrow("M049", "本月订单额", time=1, org=1, pl=1, cust=1),
    dimrow("M050", "订单环比增长率", time=1, org=1, pl=1, cust=1),
    dimrow("M051", "低毛利订单占比", time=1, pl=1, cust=1),
    dimrow("M052", "高折扣订单占比", time=1, pl=1, cust=1),
    dimrow("M053", "订单可转收入比例", time=1, pl=1),
    dimrow("M054", "订单转收入滞留金额", time=1, pl=1, status=1),
    dimrow("M055", "经营现金转化率", time=1, org=1),
    dimrow("M056", "应收净增加", time=1, cust=1),
    dimrow("M057", "逾期应收金额", time=1, cust=1),
    dimrow("M058", "重点客户应收贡献占比", cust=1),
    dimrow("M059", "净利润环比增长率", time=1, org=1),
    dimrow("M060", "价格端利润影响", time=1, mgmt=1),
    dimrow("M061", "低毛利业务利润拖累", time=1, pl=1, cust=1),
]


def last_data_row(ws):
    """真实最后一个非空数据行(按 A 列扫描),用于紧凑追加。"""
    r = ws.max_row
    while r >= 1:
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            return r
        r -= 1
    return 1


def purge(ws, code_col, prefixes):
    """删除已存在的本次相关行(幂等);从下往上删。"""
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=code_col).value
        s = "" if v is None else str(v).strip()
        if any(s == p or s.startswith(p) for p in prefixes):
            ws.delete_rows(r, 1)


def append_rows(ws, rows):
    start = last_data_row(ws) + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            ws.cell(row=start + i, column=j + 1, value=val)


def main():
    wb = openpyxl.load_workbook(XLSX)

    # ---- 幂等清理(按编码) -------------------------------------------
    purge(wb["术语"], 1, [t[0] for t in TERMS])
    purge(wb["业务规则"], 1, [r[0] for r in RULES])
    purge(wb["活动"], 1, [a[0] for a in ACTIVITIES])
    purge(wb["业务对象"], 1, [BO_CODE])
    purge(wb["逻辑实体"], 3, [t[0] for t in TABLES])      # LE 编码在第3列
    purge(wb["业务属性"], 1, [t[0] for t in TABLES])      # 按 LE 编码(第1列)清,安全
    purge(wb["实体关系"], 1, [r[0] for r in RELATIONS])
    purge(wb["指标"], 1, [m[0] for m in METRICS])
    purge(wb["指标维度矩阵"], 1, [d[0] for d in DIMROWS])

    # ---- 术语 ---------------------------------------------------------
    append_rows(wb["术语"], TERMS)

    # ---- 业务规则 -----------------------------------------------------
    append_rows(wb["业务规则"], RULES)

    # ---- 活动 ---------------------------------------------------------
    append_rows(wb["活动"], ACTIVITIES)

    # ---- 业务对象 -----------------------------------------------------
    append_rows(wb["业务对象"], [(
        BO_CODE, BO_NAME, "CFO Operating Analysis",
        "围绕订-发-收-回与盈利四个方向的 CFO 经营故事线分析域,涵盖订单质量、"
        "订单转收入卡点、收入转现金、重点客户回款、利润桥与低毛利清单。",
        "数据模型", BO_TABLES,
    )])

    # ---- 逻辑实体 + 业务属性 -----------------------------------------
    le_rows = []
    at_rows = []
    at_seq = 359
    for le_code, table, le_name, le_en, le_def, cols in TABLES:
        le_rows.append((BO_CODE, BO_NAME, le_code, le_name, le_en, le_def, table))
        for cn, fcol, dtype, cdef, pk in cols:
            at_rows.append((le_code, le_name, le_en,
                            "AT%05d" % at_seq, cn, fcol, cdef, dtype, pk, fcol))
            at_seq += 1
    append_rows(wb["逻辑实体"], le_rows)
    append_rows(wb["业务属性"], at_rows)

    # ---- 实体关系 -----------------------------------------------------
    append_rows(wb["实体关系"], RELATIONS)

    # ---- 指标 ---------------------------------------------------------
    append_rows(wb["指标"], METRICS)

    # ---- 指标维度矩阵 -------------------------------------------------
    append_rows(wb["指标维度矩阵"], DIMROWS)

    wb.save(XLSX)

    print("CFO 本体登记完成 (ChatBI业务元数据_硕磐财务管理.xlsx):")
    print("  术语        +%d (T000050~T000060)" % len(TERMS))
    print("  业务规则    +%d (R035~R041)" % len(RULES))
    print("  活动        +%d (A012)" % len(ACTIVITIES))
    print("  业务对象    +1 (%s %s)" % (BO_CODE, BO_NAME))
    print("  逻辑实体    +%d (LE00053~LE00060)" % len(le_rows))
    print("  业务属性    +%d (AT00359~AT%05d)" % (len(at_rows), at_seq - 1))
    print("  实体关系    +%d (ER105~ER113)" % len(RELATIONS))
    print("  指标        +%d (M049~M061)" % len(METRICS))
    print("  指标维度矩阵 +%d" % len(DIMROWS))


if __name__ == "__main__":
    main()
