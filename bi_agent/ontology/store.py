"""
OntologyStore — loads the ChatBI 业务元数据 xlsx into memory and provides
indexed lookup + keyword search for the BI agent.

Design:
- Single entry point: OntologyStore.from_xlsx(path) returns a fully-loaded store.
- In-memory dicts keyed by code, plus secondary indexes by name and alias.
- Lookup APIs are thin; the agent does heavier reasoning via to_prompt() snippets.
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Iterable, Optional

import openpyxl

# Some exported ontology files carry an <autoFilter> whose <customFilter> value
# openpyxl can't parse (raises "Value must be either numerical or a string
# containing a wildcard"). We only read cell values, so strip the filter XML
# in-memory before loading. Matches both self-closing and container forms.
_AUTOFILTER_RE = re.compile(
    rb"<autoFilter\b[^>]*?/>|<autoFilter\b.*?</autoFilter>", re.DOTALL
)

from .schema import (
    Activity,
    Attribute,
    BusinessObject,
    BusinessRule,
    Dimension,
    DimMatrixEntry,
    EntityRelation,
    LogicalEntity,
    METRIC_TYPE_ATOMIC,
    MetaRelation,
    Metric,
    Process,
    Term,
)


# Sheet name constants (match the real xlsx exactly)
SHEET_TERM = "术语"
SHEET_BO = "业务对象"
SHEET_LE = "逻辑实体"
SHEET_AT = "业务属性"
SHEET_ER = "实体关系"
SHEET_METRIC = "指标"
SHEET_ACTIVITY = "活动"
SHEET_RULE = "业务规则"
SHEET_DIM_MATRIX = "指标维度矩阵"
SHEET_DIM = "维度"            # 维度元素(可能缺失,旧本体文件没有此表)
SHEET_ACTIVITY_FLOW = "活动流"  # 流程(场景子流程) + 活动流转(可能缺失)
SHEET_META_REL = "本体元模型关系"  # 类型级关系(可能缺失,旧本体文件没有此表)


def _norm(v) -> str:
    """Coerce an xlsx cell to a stripped string (empty string for None)."""
    if v is None:
        return ""
    return str(v).strip()


def _split_list(v) -> list[str]:
    """Split a multi-value cell separated by comma / Chinese comma / semicolon."""
    s = _norm(v)
    if not s:
        return []
    for sep in [",", ",", ";", ";", "、"]:
        s = s.replace(sep, "|")
    return [x.strip() for x in s.split("|") if x.strip()]


def _truthy(v) -> bool:
    s = _norm(v).lower()
    return s in {"y", "yes", "1", "true", "是"}


class OntologyStore:
    """In-memory ontology loaded from the ChatBI xlsx."""

    def __init__(self) -> None:
        self.terms: dict[str, Term] = {}
        self.business_objects: dict[str, BusinessObject] = {}
        self.logical_entities: dict[str, LogicalEntity] = {}
        self.attributes: dict[str, Attribute] = {}
        self.relations: dict[str, EntityRelation] = {}
        self.meta_relations: dict[str, MetaRelation] = {}  # 类型级关系(可能为空)
        self.metrics: dict[str, Metric] = {}
        self.activities: dict[str, Activity] = {}
        self.rules: dict[str, BusinessRule] = {}
        self.dim_matrix: dict[str, DimMatrixEntry] = {}
        self.dimensions: dict[str, Dimension] = {}     # 维度元素(可能为空)
        self.processes: dict[str, Process] = {}        # 流程(场景子流程,可能为空)

        # Secondary indexes
        self._term_by_name: dict[str, str] = {}        # name/alias (lower) -> code
        self._bo_by_name: dict[str, str] = {}
        self._le_by_name: dict[str, str] = {}
        self._le_by_english: dict[str, str] = {}       # english (lower) -> le_code
        self._le_by_table: dict[str, str] = {}         # physical_table -> le_code
        self._metric_by_name: dict[str, str] = {}      # name/alias (lower) -> code
        self._dim_by_shortname: dict[str, str] = {}    # 维度矩阵列名/别名 (lower) -> dim_code
        self._attrs_by_le: dict[str, list[str]] = {}   # le_code -> [at_code, ...]
        self._les_by_bo: dict[str, list[str]] = {}     # bo_code -> [le_code, ...]
        self._relations_by_le: dict[str, list[str]] = {}  # le_code -> [er_code, ...]
        # Activity flow transitions from 活动流 (源节点->目标节点), for上下游探索.
        self._activity_flow_next: dict[str, list[str]] = {}  # activity -> [downstream act, ...]
        self._activity_flow_prev: dict[str, list[str]] = {}  # activity -> [upstream act, ...]

    # ------------------------------------------------------------------
    # Loading
    # ------------------------------------------------------------------

    @staticmethod
    def _load_workbook(path: str | Path):
        """Load an xlsx for value reading, stripping unparseable <autoFilter>
        XML in-memory first (the file on disk is never modified)."""
        with open(path, "rb") as f:
            raw = f.read()
        out = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw)) as zin, \
                zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if (
                    item.filename.startswith("xl/worksheets/")
                    and item.filename.endswith(".xml")
                ):
                    data = _AUTOFILTER_RE.sub(b"", data)
                zout.writestr(item, data)
        out.seek(0)
        return openpyxl.load_workbook(out, data_only=True, read_only=True)

    @classmethod
    def from_xlsx(cls, path: str | Path) -> "OntologyStore":
        store = cls()
        wb = cls._load_workbook(path)
        store._load_terms(wb[SHEET_TERM])
        store._load_business_objects(wb[SHEET_BO])
        store._load_logical_entities(wb[SHEET_LE])
        store._load_attributes(wb[SHEET_AT])
        store._load_relations(wb[SHEET_ER])
        if SHEET_META_REL in wb.sheetnames:
            store._load_meta_relations(wb[SHEET_META_REL])
        if SHEET_DIM in wb.sheetnames:
            store._load_dimensions(wb[SHEET_DIM])
        store._load_metrics(wb[SHEET_METRIC])
        store._load_activities(wb[SHEET_ACTIVITY])
        if SHEET_ACTIVITY_FLOW in wb.sheetnames:
            store._load_processes(wb[SHEET_ACTIVITY_FLOW])
        store._load_rules(wb[SHEET_RULE])
        store._load_dim_matrix(wb[SHEET_DIM_MATRIX])
        wb.close()
        store._wire_metric_dimensions()
        return store

    @staticmethod
    def _first_header(it) -> "tuple | None":
        """Advance the row iterator past leading blank rows; return the first
        non-empty row as the header (or None if the sheet is empty).

        Some ontology files (e.g. MetaERP 指标) put a blank spacer row above the
        real header — skipping it lets header-name mapping bind to the actual
        columns instead of a row full of None.
        """
        for row in it:
            if row is None:
                continue
            if all(c is None or _norm(c) == "" for c in row):
                continue
            return row
        return None

    @staticmethod
    def _rows(ws) -> Iterable[tuple]:
        """Yield data rows (skip header), padding short rows with None."""
        it = ws.iter_rows(values_only=True)
        header = OntologyStore._first_header(it)
        if header is None:
            return
        width = len(header)
        for row in it:
            if row is None:
                continue
            if all(c is None or _norm(c) == "" for c in row):
                continue
            if len(row) < width:
                row = row + (None,) * (width - len(row))
            yield row

    # --- Header-driven loading -------------------------------------------
    # Some sheets (指标 / 实体关系) differ in column layout between ontology
    # files (e.g. ChatBI 17-col 指标 vs 超聚变 8-col 指标). Loaders for those
    # sheets map by header NAME instead of fixed position.

    @staticmethod
    def _norm_key(s) -> str:
        """Normalize a header / lookup name: drop brackets, commas, spaces."""
        out = _norm(s)
        for ch in "（）()【】[]，,、 　":
            out = out.replace(ch, "")
        return out

    @staticmethod
    def _col_index(header) -> dict:
        """Map normalized header names -> column index (first occurrence wins)."""
        idx: dict[str, int] = {}
        for i, h in enumerate(header or ()):
            key = OntologyStore._norm_key(h)
            if key and key not in idx:
                idx[key] = i
        return idx

    @staticmethod
    def _cell(row, col_idx: dict, *names: str) -> str:
        """Read a row value by any candidate header name (exact, then prefix)."""
        for n in names:
            key = OntologyStore._norm_key(n)
            if not key:
                continue
            pos = col_idx.get(key)
            if pos is None:
                for hk, hp in col_idx.items():
                    if hk.startswith(key):
                        pos = hp
                        break
            if pos is not None and pos < len(row):
                return _norm(row[pos])
        return ""

    @staticmethod
    def _indexed(ws):
        """Return (col_index, data_rows) for a sheet — header mapped by name."""
        it = ws.iter_rows(values_only=True)
        header = OntologyStore._first_header(it)
        col = OntologyStore._col_index(header)
        rows = [
            row for row in it
            if row is not None and not all(c is None or _norm(c) == "" for c in row)
        ]
        return col, rows

    def _load_terms(self, ws) -> None:
        # Header-driven: column order/extra columns vary across ontology files
        # (e.g. MetaERP 术语 inserts a 缩略语 column that shifts定义/分类).
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "术语编码", "术语编号")
            if not code:
                continue
            term = Term(
                code=code,
                name=self._cell(row, col, "术语名称"),
                aliases=_split_list(self._cell(row, col, "别名")),
                english=self._cell(row, col, "术语英文名", "英文名"),
                definition=self._cell(row, col, "术语定义"),
                category=self._cell(row, col, "术语分类"),
                department=self._cell(row, col, "解释部门", "部门"),
            )
            self.terms[code] = term
            if term.name:
                self._term_by_name[term.name.lower()] = code
            for alias in term.aliases:
                self._term_by_name[alias.lower()] = code

    def _load_business_objects(self, ws) -> None:
        # Header-driven: MetaERP 业务对象 omits 类型/物理表 entirely.
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "业务对象编号", "业务对象编码")
            if not code:
                continue
            bo = BusinessObject(
                code=code,
                name=self._cell(row, col, "业务对象名称"),
                english=self._cell(row, col, "业务对象英文名"),
                definition=self._cell(row, col, "业务对象定义"),
                type=self._cell(row, col, "业务对象类型"),
                physical_table=self._cell(row, col, "关联物理表", "对应物理表"),
            )
            self.business_objects[code] = bo
            if bo.name:
                self._bo_by_name[bo.name.lower()] = code

    def _load_logical_entities(self, ws) -> None:
        # Header-driven: MetaERP 逻辑实体 has no 物理表 column (语义层用英文名).
        col, rows = self._indexed(ws)
        for row in rows:
            bo_code = self._cell(row, col, "业务对象编号", "业务对象编码")
            bo_name = self._cell(row, col, "业务对象名称")
            le_code = self._cell(row, col, "逻辑实体编号", "逻辑实体编码")
            if not le_code:
                continue
            le = LogicalEntity(
                code=le_code,
                name=self._cell(row, col, "逻辑实体名称"),
                english=self._cell(row, col, "英文名", "逻辑实体英文名"),
                definition=self._cell(row, col, "逻辑实体定义"),
                physical_table=self._cell(row, col, "对应物理表", "关联物理表"),
                bo_code=bo_code,
                bo_name=bo_name,
            )
            self.logical_entities[le_code] = le
            if le.name:
                self._le_by_name[le.name.lower()] = le_code
            if le.english:
                self._le_by_english[le.english.lower()] = le_code
            if le.physical_table:
                self._le_by_table[le.physical_table] = le_code
            if bo_code:
                self._les_by_bo.setdefault(bo_code, []).append(le_code)

    def _load_attributes(self, ws) -> None:
        # Header-driven: MetaERP 业务属性 has no 逻辑实体编号 column and splits
        # 是否主键 into 是否物理主键/逻辑主键 — positional mapping mis-aligns.
        col, rows = self._indexed(ws)
        for row in rows:
            at_code = self._cell(row, col, "业务属性编号", "业务属性编码", "属性编号")
            if not at_code:
                continue
            le_name = self._cell(row, col, "逻辑实体名称")
            le_code = self._cell(row, col, "逻辑实体编号", "逻辑实体编码")
            # MetaERP 业务属性 links to its entity by NAME (no code column);
            # backfill the code from the LE name so drill-down (_attrs_by_le)
            # still works. LE is loaded before AT, so _le_by_name is ready.
            if not le_code and le_name:
                le_code = self._le_by_name.get(le_name.lower(), "")
            attr = Attribute(
                code=at_code,
                name=self._cell(row, col, "业务属性名称", "属性名称"),
                english=self._cell(row, col, "业务属性英文名称", "业务属性英文名", "属性英文名"),
                definition=self._cell(row, col, "业务属性定义", "属性定义"),
                data_type=self._cell(row, col, "数据类型"),
                is_primary_key=_truthy(self._cell(row, col, "是否物理主键", "是否主键")),
                le_code=le_code,
                le_name=le_name,
                le_english=self._cell(row, col, "逻辑实体英文名"),
            )
            self.attributes[at_code] = attr
            if le_code:
                self._attrs_by_le.setdefault(le_code, []).append(at_code)

    def _load_relations(self, ws) -> None:
        # Header-driven: ChatBI 源/目标实体, 超聚变 主/从实体, MetaERP 源/目标逻辑实体.
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "关系编号")
            if not code:
                continue
            er = EntityRelation(
                code=code,
                source_code=self._cell(row, col, "源实体编号", "主实体编号", "源逻辑实体编号"),
                source_name=self._cell(row, col, "源实体名称", "主实体名称", "源逻辑实体名称"),
                target_code=self._cell(row, col, "目标实体编号", "从实体编号", "目标逻辑实体编号"),
                target_name=self._cell(row, col, "目标实体名称", "从实体名称", "目标逻辑实体名称"),
                cardinality=self._cell(row, col, "关系类型", "关系基数"),
                description=self._cell(row, col, "关系描述", "关系说明"),
                foreign_key=self._cell(row, col, "外键字段", "关联条件", "源关联属性英文名"),
                name=self._cell(row, col, "关系中文名称", "关系名称"),
                english=self._cell(row, col, "关系英文名称"),
            )
            self.relations[code] = er
            if er.source_code:
                self._relations_by_le.setdefault(er.source_code, []).append(code)
            if er.target_code and er.target_code != er.source_code:
                self._relations_by_le.setdefault(er.target_code, []).append(code)

    def _load_meta_relations(self, ws) -> None:
        # 本体元模型关系 — 类型级关系(源/目标是本体元素类型 token,如
        # business_object / logical_entity)。旧本体文件没有此表,from_xlsx 已守卫。
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "关系编号")
            if not code:
                continue
            mr = MetaRelation(
                code=code,
                source_kind=self._cell(row, col, "源本体元素编码", "源元素编码", "源本体元素"),
                source_name=self._cell(row, col, "源本体元素名称", "源元素名称"),
                target_kind=self._cell(row, col, "目标本体元素编码", "目标元素编码", "目标本体元素"),
                target_name=self._cell(row, col, "目标本体元素名称", "目标元素名称"),
                category=self._cell(row, col, "关系分类", "关系类型"),
                name=self._cell(row, col, "关系中文名称", "关系名称"),
                english=self._cell(row, col, "关系英文名称"),
                cardinality=self._cell(row, col, "关系基数"),
                description=self._cell(row, col, "关系描述"),
                reverse_name=self._cell(row, col, "反向关系中文名称"),
                reverse_english=self._cell(row, col, "反向关系英文名称"),
                agent_use=_truthy(self._cell(row, col, "是否智能体使用")),
            )
            self.meta_relations[code] = mr

    def _load_dimensions(self, ws) -> None:
        # 维度 — 指标下钻的分析轴。指标维度矩阵的列名是维度短名(时间/管理单元…),
        # 维度页用全名(时间维度/管理单元维度…);建立短名->编码索引以便对齐。
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "维度编码", "维度编号")
            if not code:
                continue
            aliases = _split_list(self._cell(row, col, "维度别名"))
            dim = Dimension(
                code=code,
                name=self._cell(row, col, "维度名称"),
                aliases=aliases,
                english=self._cell(row, col, "维度英文名"),
                definition=self._cell(row, col, "维度定义"),
                level=self._cell(row, col, "维度层级"),
            )
            self.dimensions[code] = dim
            # index by full name, by name without 维度 suffix, and by aliases
            keys = {dim.name, dim.name.replace("维度", "").strip(), *aliases}
            for k in keys:
                if k:
                    self._dim_by_shortname[k.lower()] = code

    def _load_processes(self, ws) -> None:
        # 活动流 — 每行是一条活动流转(源节点->目标节点)。流程(场景子流程)
        # 由 场景子流程编码/所属场景子流程 标识,其包含的活动 = 出现在该流程
        # 各流转行里的源/目标活动(去重保序)。
        col, rows = self._indexed(ws)
        for row in rows:
            pcode = self._cell(row, col, "场景子流程编码", "流程编码")
            pname = self._cell(row, col, "所属场景子流程", "流程名称")
            if not pcode:
                continue
            proc = self.processes.get(pcode)
            if proc is None:
                proc = Process(code=pcode, name=pname)
                self.processes[pcode] = proc
            src = self._cell(row, col, "源节点编号", "源活动编号")
            tgt = self._cell(row, col, "目标节点编号", "目标活动编号")
            for act in (src, tgt):
                if act and act in self.activities and act not in proc.activity_codes:
                    proc.activity_codes.append(act)
            # Flow transition src -> tgt (both must be known activities).
            if src in self.activities and tgt in self.activities:
                nxt = self._activity_flow_next.setdefault(src, [])
                if tgt not in nxt:
                    nxt.append(tgt)
                prv = self._activity_flow_prev.setdefault(tgt, [])
                if src not in prv:
                    prv.append(src)

    def _load_metrics(self, ws) -> None:
        # Header-driven: ChatBI 指标 has 17 rich columns; 超聚变 指标 has 8
        # (编号/名称/英文名/定义/计算口径/数据类型/来源实体/解释部门).
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "指标编码", "指标编号")
            if not code:
                continue
            type_raw = self._cell(row, col, "指标类型")
            try:
                mtype = int(type_raw) if type_raw else METRIC_TYPE_ATOMIC
            except (TypeError, ValueError):
                mtype = METRIC_TYPE_ATOMIC
            # 超聚变 carries a single 计算口径 column; reuse it where the
            # richer ChatBI 业务公式 / 统计口径 columns are absent.
            caliber = self._cell(row, col, "计算口径")
            metric = Metric(
                code=code,
                name=self._cell(row, col, "指标名称"),
                aliases=_split_list(self._cell(row, col, "指标别名")),
                english=self._cell(row, col, "指标英文名"),
                definition=self._cell(row, col, "指标定义"),
                formula_business=self._cell(row, col, "计算公式(业务)") or caliber,
                scope=self._cell(row, col, "统计口径") or caliber,
                metric_type=mtype,
                formula_technical=self._cell(row, col, "计算公式(技术)"),
                table=self._cell(row, col, "对应表名", "对应表名/逻辑实体", "来源实体"),
                bo_name=self._cell(row, col, "业务对象"),
                agg_column=self._cell(row, col, "聚合列", "聚合列/属性"),
                agg_type=self._cell(row, col, "聚合类型"),
                join_condition=self._cell(row, col, "连接条件"),
                filter_condition=self._cell(row, col, "筛选条件"),
                dim_columns=self._cell(row, col, "维度列"),
                version=self._cell(row, col, "版本"),
                change_note=self._cell(row, col, "修改说明"),
            )
            self.metrics[code] = metric
            if metric.name:
                self._metric_by_name[metric.name.lower()] = code
            for alias in metric.aliases:
                self._metric_by_name[alias.lower()] = code

    def _load_activities(self, ws) -> None:
        # Header-driven: MetaERP 活动 omits 上游活动/业务角色 and reorders cols.
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "活动编码", "活动编号")
            if not code:
                continue
            act = Activity(
                code=code,
                name=self._cell(row, col, "活动名称"),
                description=self._cell(row, col, "活动描述", "活动定义"),
                upstream_codes=_split_list(self._cell(row, col, "上游活动编码", "上游活动编号")),
                role=self._cell(row, col, "业务角色"),
                object_codes=_split_list(self._cell(row, col, "操作业务对象编号")),
                object_names=_split_list(self._cell(row, col, "操作业务对象名称")),
            )
            self.activities[code] = act

    def _load_rules(self, ws) -> None:
        # Header-driven: MetaERP 业务规则 inserts 规则描述, shifting cols.
        col, rows = self._indexed(ws)
        for row in rows:
            code = self._cell(row, col, "规则编号", "规则编码")
            if not code:
                continue
            rule = BusinessRule(
                code=code,
                rule=self._cell(row, col, "规则名称", "规则"),
                activity_code=self._cell(row, col, "所属活动编号", "所属活动编码"),
                activity_name=self._cell(row, col, "活动名称"),
                category=self._cell(row, col, "规则分类"),
                department=self._cell(row, col, "解释部门", "部门"),
            )
            self.rules[code] = rule

    def _load_dim_matrix(self, ws) -> None:
        it = ws.iter_rows(values_only=True)
        header = self._first_header(it)
        if header is None:
            return
        dim_names = [_norm(h) for h in header[2:]]  # cols 0,1 are metric code/name
        for row in it:
            if row is None or not _norm(row[0]):
                continue
            metric_code = _norm(row[0])
            metric_name = _norm(row[1])
            dims: dict[str, bool] = {}
            for i, dim in enumerate(dim_names):
                if not dim:
                    continue
                cell = row[2 + i] if 2 + i < len(row) else None
                dims[dim] = _truthy(cell)
            self.dim_matrix[metric_code] = DimMatrixEntry(
                metric_code=metric_code,
                metric_name=metric_name,
                dimensions=dims,
            )

    def _wire_metric_dimensions(self) -> None:
        """Copy applicable dimensions from dim_matrix into each Metric."""
        for code, metric in self.metrics.items():
            entry = self.dim_matrix.get(code)
            if entry:
                metric.applicable_dimensions = entry.applicable()

    # ------------------------------------------------------------------
    # Lookup APIs
    # ------------------------------------------------------------------

    def find_term(self, query: str) -> list[Term]:
        """Resolve a term by name or alias; returns all that match (usually 1)."""
        q = query.strip().lower()
        if not q:
            return []
        hits: list[Term] = []
        # Exact match on name/alias index
        code = self._term_by_name.get(q)
        if code:
            hits.append(self.terms[code])
        # Substring fallback over name/alias/definition
        if not hits:
            for t in self.terms.values():
                hay = " ".join([t.name, *t.aliases, t.english, t.definition]).lower()
                if q in hay:
                    hits.append(t)
        return hits

    def find_metric(self, query: str) -> list[Metric]:
        q = query.strip().lower()
        if not q:
            return []
        code = self._metric_by_name.get(q)
        if code:
            return [self.metrics[code]]
        hits: list[Metric] = []
        for m in self.metrics.values():
            hay = " ".join([m.name, *m.aliases, m.english, m.definition]).lower()
            if q in hay:
                hits.append(m)
        return hits

    def get_business_object(self, code_or_name: str) -> Optional[BusinessObject]:
        key = code_or_name.strip()
        if key in self.business_objects:
            return self.business_objects[key]
        code = self._bo_by_name.get(key.lower())
        return self.business_objects.get(code) if code else None

    def get_logical_entity(self, code_or_name: str) -> Optional[LogicalEntity]:
        key = code_or_name.strip()
        if key in self.logical_entities:
            return self.logical_entities[key]
        code = self._le_by_name.get(key.lower())
        if code:
            return self.logical_entities[code]
        code = self._le_by_table.get(key)
        return self.logical_entities.get(code) if code else None

    def attributes_of(self, le_code: str) -> list[Attribute]:
        return [self.attributes[c] for c in self._attrs_by_le.get(le_code, [])]

    def logical_entities_of(self, bo_code: str) -> list[LogicalEntity]:
        return [self.logical_entities[c] for c in self._les_by_bo.get(bo_code, [])]

    def relations_of(self, le_code: str) -> list[EntityRelation]:
        seen: set[str] = set()
        out: list[EntityRelation] = []
        for c in self._relations_by_le.get(le_code, []):
            if c in seen:
                continue
            seen.add(c)
            out.append(self.relations[c])
        return out

    def rules_of(self, activity_code: str) -> list[BusinessRule]:
        return [r for r in self.rules.values() if r.activity_code == activity_code]

    def search(self, query: str, limit: int = 20) -> dict[str, list]:
        """Cross-ontology keyword search. Returns buckets by type."""
        q = query.strip().lower()
        result: dict[str, list] = {
            "terms": [], "business_objects": [], "logical_entities": [],
            "metrics": [], "activities": [], "rules": [],
        }
        if not q:
            return result

        def _match(*texts: str) -> bool:
            return q in " ".join(t for t in texts if t).lower()

        for t in self.terms.values():
            if _match(t.name, *t.aliases, t.english, t.definition):
                result["terms"].append(t)
        for bo in self.business_objects.values():
            if _match(bo.name, bo.english, bo.definition, bo.physical_table):
                result["business_objects"].append(bo)
        for le in self.logical_entities.values():
            if _match(le.name, le.english, le.definition, le.physical_table):
                result["logical_entities"].append(le)
        for m in self.metrics.values():
            if _match(m.name, *m.aliases, m.english, m.definition, m.scope):
                result["metrics"].append(m)
        for a in self.activities.values():
            if _match(a.name, a.description):
                result["activities"].append(a)
        for r in self.rules.values():
            if _match(r.rule, r.activity_name, r.category):
                result["rules"].append(r)

        for k in result:
            result[k] = result[k][:limit]
        return result

    def stats(self) -> dict[str, int]:
        return {
            "terms": len(self.terms),
            "business_objects": len(self.business_objects),
            "logical_entities": len(self.logical_entities),
            "attributes": len(self.attributes),
            "relations": len(self.relations),
            "metrics": len(self.metrics),
            "activities": len(self.activities),
            "rules": len(self.rules),
            "dimensions": len(self.dimensions),
            "processes": len(self.processes),
        }
